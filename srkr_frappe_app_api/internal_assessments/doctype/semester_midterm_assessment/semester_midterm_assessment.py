# Copyright (c) 2024, your_name and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
import openpyxl
from io import BytesIO
import os

class SemesterMidtermAssessment(Document):
    def before_save(self):
        for row in self.get("assessment_structure"):
            row.descriptive_name = f"{row.midterm}-{row.assessment_type}-{row.question_id}"
        self.recalculate_scores()

    def recalculate_scores(self):
        # ... (this function is correct and unchanged) ...
        student_totals = {row.student: {"mid_1_total": 0, "mid_2_total": 0} for row in self.final_scores_summary}
        structure_midterm_map = {row.name: row.midterm for row in self.assessment_structure}
        for marks_row in self.student_marks_data:
            midterm = structure_midterm_map.get(marks_row.assessment_item)
            student = marks_row.student
            marks = marks_row.marks_obtained or 0
            if student in student_totals:
                if midterm == "Mid-1":
                    student_totals[student]["mid_1_total"] += marks
                elif midterm == "Mid-2":
                    student_totals[student]["mid_2_total"] += marks
        for summary_row in self.final_scores_summary:
            totals = student_totals.get(summary_row.student)
            
            # If Manual Entry Mode is OFF, overwrite the totals with the calculated sum.
            # If it is ON, we preserve whatever the user typed in the grid.
            if not self.manual_entry_mode and totals:
                summary_row.mid_1_total = totals["mid_1_total"]
                summary_row.mid_2_total = totals["mid_2_total"]

            # Always calculate the final weightage based on whatever is in the columns now.
            mid1, mid2 = (summary_row.mid_1_total or 0), (summary_row.mid_2_total or 0)
            higher_score, lower_score = max(mid1, mid2), min(mid1, mid2)
            final_marks = round((0.8 * higher_score) + (0.2 * lower_score))
            summary_row.total_internal_marks = final_marks

# --- WHITELISTED FUNCTIONS ---

@frappe.whitelist()
def get_students_for_group(student_group):
    # ... (this function is correct and unchanged) ...
    if not student_group: return []
    student_group_students = frappe.get_all("Student Group Student", filters={"parent": student_group}, fields=["student"])
    if not student_group_students: return []
    student_ids = [s.student for s in student_group_students]
    student_details = frappe.get_all("Student", filters={"name": ["in", student_ids], "enabled": 1}, fields=["name as student", "student_name", "custom_student_id as register_number"], order_by="custom_student_id")
    return student_details

# --- THIS IS THE FUNCTION THAT WAS ACCIDENTALLY DELETED. IT IS NOW BACK. ---
@frappe.whitelist()
def get_default_assessment_structure():
    """Returns the hardcoded default assessment structure as a list of dicts."""
    return [
        {'midterm': 'Mid-1', 'assessment_type': 'Objective', 'question_id': 'M1-Obj-Q1', 'co_number': 'CO1', 'max_marks': 2},
        # ... (and all the other 25 rows of the default structure) ...
        {'midterm': 'Mid-1', 'assessment_type': 'Objective', 'question_id': 'M1-Obj-Q2', 'co_number': 'CO1', 'max_marks': 2},
        {'midterm': 'Mid-1', 'assessment_type': 'Objective', 'question_id': 'M1-Obj-Q3', 'co_number': 'CO2', 'max_marks': 2},
        {'midterm': 'Mid-1', 'assessment_type': 'Objective', 'question_id': 'M1-Obj-Q4', 'co_number': 'CO2', 'max_marks': 2},
        {'midterm': 'Mid-1', 'assessment_type': 'Objective', 'question_id': 'M1-Obj-Q5', 'co_number': 'CO3', 'max_marks': 2},
        {'midterm': 'Mid-1', 'assessment_type': 'Subjective', 'question_id': 'M1-Subj-Q1', 'co_number': 'CO1', 'max_marks': 5},
        {'midterm': 'Mid-1', 'assessment_type': 'Subjective', 'question_id': 'M1-Subj-Q2', 'co_number': 'CO2', 'max_marks': 5},
        {'midterm': 'Mid-1', 'assessment_type': 'Subjective', 'question_id': 'M1-Subj-Q3', 'co_number': 'CO3', 'max_marks': 5},
        {'midterm': 'Mid-1', 'assessment_type': 'Assignment', 'question_id': 'M1-Asgn-Q1', 'co_number': 'CO1', 'max_marks': 3},
        {'midterm': 'Mid-1', 'assessment_type': 'Assignment', 'question_id': 'M1-Asgn-Q2', 'co_number': 'CO1', 'max_marks': 3},
        {'midterm': 'Mid-1', 'assessment_type': 'Assignment', 'question_id': 'M1-Asgn-Q3', 'co_number': 'CO2', 'max_marks': 3},
        {'midterm': 'Mid-1', 'assessment_type': 'Assignment', 'question_id': 'M1-Asgn-Q4', 'co_number': 'CO2', 'max_marks': 3},
        {'midterm': 'Mid-1', 'assessment_type': 'Assignment', 'question_id': 'M1-Asgn-Q5', 'co_number': 'CO3', 'max_marks': 3},
        {'midterm': 'Mid-2', 'assessment_type': 'Objective', 'question_id': 'M2-Obj-Q1', 'co_number': 'CO4', 'max_marks': 2},
        {'midterm': 'Mid-2', 'assessment_type': 'Objective', 'question_id': 'M2-Obj-Q2', 'co_number': 'CO4', 'max_marks': 2},
        {'midterm': 'Mid-2', 'assessment_type': 'Objective', 'question_id': 'M2-Obj-Q3', 'co_number': 'CO5', 'max_marks': 2},
        {'midterm': 'Mid-2', 'assessment_type': 'Objective', 'question_id': 'M2-Obj-Q4', 'co_number': 'CO5', 'max_marks': 2},
        {'midterm': 'Mid-2', 'assessment_type': 'Objective', 'question_id': 'M2-Obj-Q5', 'co_number': 'CO5', 'max_marks': 2},
        {'midterm': 'Mid-2', 'assessment_type': 'Subjective', 'question_id': 'M2-Subj-Q1', 'co_number': 'CO4', 'max_marks': 5},
        {'midterm': 'Mid-2', 'assessment_type': 'Subjective', 'question_id': 'M2-Subj-Q2', 'co_number': 'CO5', 'max_marks': 5},
        {'midterm': 'Mid-2', 'assessment_type': 'Subjective', 'question_id': 'M2-Subj-Q3', 'co_number': 'CO5', 'max_marks': 5},
        {'midterm': 'Mid-2', 'assessment_type': 'Assignment', 'question_id': 'M2-Asgn-Q1', 'co_number': 'CO4', 'max_marks': 3},
        {'midterm': 'Mid-2', 'assessment_type': 'Assignment', 'question_id': 'M2-Asgn-Q2', 'co_number': 'CO4', 'max_marks': 3},
        {'midterm': 'Mid-2', 'assessment_type': 'Assignment', 'question_id': 'M2-Asgn-Q3', 'co_number': 'CO5', 'max_marks': 3},
        {'midterm': 'Mid-2', 'assessment_type': 'Assignment', 'question_id': 'M2-Asgn-Q4', 'co_number': 'CO5', 'max_marks': 3},
        {'midterm': 'Mid-2', 'assessment_type': 'Assignment', 'question_id': 'M2-Asgn-Q5', 'co_number': 'CO5', 'max_marks': 3},
    ]

@frappe.whitelist()
def generate_marksheet_template(docname):
    # ... (this function is correct and unchanged) ...
    doc = frappe.get_doc("Semester Midterm Assessment", docname)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Marksheet"
    headers = ["Register Number", "Student Name"]
    
    # Dynamic Template Generation
    if doc.manual_entry_mode:
        # In Manual Mode, we only need the totals.
        descriptive_headers = ["Mid-1 Total", "Mid-2 Total"]
    else:
        # In Standard Mode, we need the full breakdown.
        descriptive_headers = [f"{row.midterm}-{row.assessment_type}-{row.question_id}" for row in doc.assessment_structure]
    
    headers.extend(descriptive_headers)
    sheet.append(headers)
    for student_row in doc.final_scores_summary:
        sheet.append([student_row.customer_student_id, student_row.student_name])
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    frappe.local.response.filename = f"{doc.course}-{doc.student_group}-Marksheet-Template.xlsx"
    frappe.local.response.filecontent = excel_file.read()
    frappe.local.response.type = "binary"


@frappe.whitelist()
def upload_marksheet(file_url, docname):
    doc = frappe.get_doc("Semester Midterm Assessment", docname)

    # --- S3-COMPATIBLE FILE READING LOGIC ---

    # Step 1: Find the File document using the URL provided by the uploader.
    file_doc = frappe.get_doc("File", {"file_url": file_url})
    if not file_doc:
        frappe.throw(f"Could not find the uploaded file record for URL: {file_url}")

    # Step 2: Get the raw content. The .get_content() method is S3-aware.
    content = file_doc.get_content() # This will return either a string or bytes

    # Step 3: Ensure we have bytes for the openpyxl library.
    if isinstance(content, str):
        # If Frappe gave us a decoded string, encode it back to raw bytes.
        content_bytes = content.encode('utf-8')
    else:
        # If Frappe already gave us bytes, use them directly.
        content_bytes = content

    # Step 4: Load the workbook from the in-memory bytes.
    workbook = openpyxl.load_workbook(filename=BytesIO(content_bytes), data_only=True)
    # --- END OF S3-COMPATIBLE LOGIC ---
    
    sheet = workbook.active
    header = [cell.value for cell in sheet[1]]
    
    # Create maps for faster lookups during processing.
    student_map = {row.customer_student_id: row.student for row in doc.final_scores_summary}
    structure_map = {f"{row.midterm}-{row.assessment_type}-{row.question_id}": row.name for row in doc.assessment_structure}
    
    # Clear any old data before importing the new marks, BUT ONLY IF IN STANDARD MODE.
    # In Manual Mode, we want to preserve the underlying granular data.
    if not doc.manual_entry_mode:
        doc.student_marks_data = []
    
    # Loop through the Excel rows, starting from the second row (to skip the header).
    for row_index in range(2, sheet.max_row + 1):
        register_number = sheet.cell(row=row_index, column=1).value
        if not register_number: 
            continue # Skip empty rows

        student_docname = student_map.get(str(register_number).strip())
        if not student_docname: 
            # This student is in the Excel file but not in the official roster. Skip them.
            continue

        # Loop through the question columns in the Excel file.
        for col_index in range(3, len(header) + 1):
            descriptive_header_from_excel = header[col_index - 1]
            marks = sheet.cell(row=row_index, column=col_index).value
            
            # --- MANUAL ENTRY MODE LOGIC ---
            if doc.manual_entry_mode:
                # In Manual Mode, we expect headers "Mid-1 Total" and "Mid-2 Total"
                # We update the summary table directly.
                summary_row = next((r for r in doc.final_scores_summary if r.customer_student_id == str(register_number).strip()), None)
                if summary_row:
                    if descriptive_header_from_excel == "Mid-1 Total":
                        summary_row.mid_1_total = marks if marks is not None else 0
                    elif descriptive_header_from_excel == "Mid-2 Total":
                        summary_row.mid_2_total = marks if marks is not None else 0
            
            # --- STANDARD MODE LOGIC ---
            else:
                # Look up the internal name of the assessment item using the descriptive header.
                assessment_item_docname = structure_map.get(descriptive_header_from_excel)
                
                if assessment_item_docname:
                    # Add a new row to the hidden "Student Marks Data" table.
                    doc.append("student_marks_data", {
                        "student": student_docname,
                        "assessment_item": assessment_item_docname,
                        "marks_obtained": marks if marks is not None else 0
                    })
    
    # Save the document. This will trigger the `before_save` hook, which runs the
    # `recalculate_scores` function to update the final summary table.
    doc.save(ignore_permissions=True)
    frappe.db.commit()
    
    return doc

@frappe.whitelist()
def trigger_recalculation(docname):
    doc = frappe.get_doc("Semester Midterm Assessment", docname)
    doc.recalculate_scores()
    doc.save()
    return "Recalculation Complete"